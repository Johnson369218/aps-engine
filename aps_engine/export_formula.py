# -*- coding: utf-8 -*-
"""通用排产表生成器 v2（5-Sheet 公式联动·可核查）——APS Engine 标准输出契约。"""
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ['产线', '订单号', '产品ID', '产品名称', '数量', '交期', '开始', '结束',
           '生产日期', '延期(分)', '准时(1/0)', '负荷(小时)', '延期(天)']

def _rhup(x, nd=2):
    return float(Decimal(str(x)).quantize(Decimal('1.' + '0' * nd), rounding=ROUND_HALF_UP))

def _dt(v):
    for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(str(v), fmt)
        except ValueError:
            continue
    return None

def export_formula_xlsx(result, products, path, title=None, assumptions=None, sensitivity=None):
    prod_by_id = {p['id']: p for p in products}
    checks = []
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill('solid', fgColor='2F5597')
    hfont = Font(bold=True, color='FFFFFF')
    tfont = Font(bold=True, size=13)
    # 主数据
    ws_m = wb.create_sheet('主数据')
    for c, h in enumerate(['产品ID', '产品名称', '单位', '8h产能'], start=1):
        cell = ws_m.cell(row=1, column=c, value=h)
        cell.fill, cell.font, cell.border = hfill, hfont, border
    for i, p in enumerate(sorted(products, key=lambda x: x['id']), start=2):
        cap = p.get('capacity_8h') or (p.get('speed_per_hour') or 0) * 8
        ws_m.cell(row=i, column=1, value=p['id'])
        ws_m.cell(row=i, column=2, value=p.get('name', p['id']))
        ws_m.cell(row=i, column=3, value=p.get('unit', ''))
        ws_m.cell(row=i, column=4, value=cap)
    for col in 'ABCD':
        ws_m.column_dimensions[col].width = 18
    ws_m.freeze_panes = 'A2'
    # 明细
    ws = wb.active
    ws.title = '排产明细'
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font, cell.border = hfill, hfont, border
        cell.alignment = Alignment(horizontal='center')
    r = 2
    for blk in result['schedule']:
        for t in blk['tasks']:
            end_dt = _dt(t['end'])
            ws.cell(row=r, column=1, value=blk['line'])
            ws.cell(row=r, column=2, value=t['order'])
            ws.cell(row=r, column=3, value=t['product'])
            ws.cell(row=r, column=4, value=t.get('product_name', t['product']))
            ws.cell(row=r, column=5, value=t['qty'])
            ws.cell(row=r, column=6, value=_dt(t['due']))
            ws.cell(row=r, column=7, value=_dt(t['start']))
            ws.cell(row=r, column=8, value=end_dt)
            ws.cell(row=r, column=9, value=end_dt.strftime('%Y-%m-%d') if end_dt else t.get('prod_date', ''))
            ws.cell(row=r, column=10, value='=MAX(0,ROUND((H{0}-F{0})*1440,0))'.format(r))
            ws.cell(row=r, column=11, value='=IF(J{0}=0,1,0)'.format(r))
            ws.cell(row=r, column=12, value='=IFERROR(ROUND(E{0}/VLOOKUP(C{0},主数据!$A:$D,4,0)*8,2),0)'.format(r))
            ws.cell(row=r, column=13, value='=IF(J{0}=0,0,ROUND(J{0}/1440,1))'.format(r))
            for c in range(1, 14):
                ws.cell(row=r, column=c).border = border
            due_dt = _dt(t['due'])
            exp_tardy = max(0, int((end_dt - due_dt).total_seconds() / 60)) if (end_dt and due_dt) else 0
            p = prod_by_id.get(t['product'], {})
            cap = p.get('capacity_8h') or (p.get('speed_per_hour') or 0) * 8
            exp_load = _rhup(t['qty'] / cap * 8) if cap else 0
            checks.append(('J', r, exp_tardy))
            checks.append(('L', r, exp_load))
            r += 1
    last = r - 1
    for i, w in enumerate([8, 26, 14, 26, 8, 18, 18, 18, 12, 10, 10, 10, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:M{0}'.format(last)
    # 汇总（公式勾稽）
    ws_s = wb.create_sheet('汇总')
    ws_s.cell(row=1, column=1, value=(title or 'APS Engine 排产汇总') + '（公式联动，打开自动重算）').font = tfont
    s = result['summary']
    ws_s.cell(row=2, column=1, value='生成 {0} | 引擎 {1} | 订单 {2}'.format(result.get('generated_at', ''), result.get('engine', ''), s.get('orders', '')))
    kpi = [
        ('总订单数', '=COUNTA(排产明细!$B$2:$B{0})'.format(last)),
        ('准时数', '=COUNTIF(排产明细!$K$2:$K{0},1)'.format(last)),
        ('延期数', '=B4-B5'),
        ('准时率', '=IFERROR(B5/B4,0)'),
        ('总延期(分钟)', '=SUM(排产明细!$J$2:$J{0})'.format(last)),
        ('最大延期(分钟)', '=MAX(排产明细!$J$2:$J{0})'.format(last)),
        ('总负荷(小时)', '=SUM(排产明细!$L$2:$L{0})'.format(last)),
    ]
    ws_s.cell(row=3, column=1, value='指标').font = Font(bold=True)
    ws_s.cell(row=3, column=2, value='公式').font = Font(bold=True)
    ws_s.cell(row=3, column=3, value='口径').font = Font(bold=True)
    for i, (name, fml) in enumerate(kpi, start=4):
        ws_s.cell(row=i, column=1, value=name)
        ws_s.cell(row=i, column=2, value=fml)
        ws_s.cell(row=i, column=3, value='勾稽明细公式')
    row0 = len(kpi) + 5
    for c, h in enumerate(['产线', '任务数', '总延期(分)', '准时数', '准时率'], start=1):
        ws_s.cell(row=row0, column=c, value=h).font = Font(bold=True)
    for i, ln in enumerate(sorted({blk['line'] for blk in result['schedule']}), start=row0 + 1):
        ws_s.cell(row=i, column=1, value=ln)
        ws_s.cell(row=i, column=2, value='=COUNTIF(排产明细!$A$2:$A{0},$A{1})'.format(last, i))
        ws_s.cell(row=i, column=3, value='=SUMIF(排产明细!$A$2:$A{0},$A{1},排产明细!$J$2:$J{0})'.format(last, i))
        ws_s.cell(row=i, column=4, value='=COUNTIFS(排产明细!$A$2:$A{0},$A{1},排产明细!$K$2:$K{0},1)'.format(last, i))
        ws_s.cell(row=i, column=5, value='=IFERROR(D{0}/B{0},0)'.format(i))
    for col in 'ABCDE':
        ws_s.column_dimensions[col].width = 16
    # 甘特图
    ws_g = wb.create_sheet('甘特图')
    dates = set()
    for blk in result['schedule']:
        for t in blk['tasks']:
            d = t.get('prod_date', '')
            if not d:
                e = _dt(t['end'])
                d = e.strftime('%Y-%m-%d') if e else ''
            if d:
                dates.add(d)
    dates = sorted(dates)
    ws_g.cell(row=1, column=1, value='产线 × 日期 负荷矩阵（SUMIFS 引用明细；红>8h 黄4-8h 绿<4h）').font = tfont
    hdr = 3
    ws_g.cell(row=hdr, column=1, value='产线 / 日期').font = Font(bold=True)
    for i, d in enumerate(dates, start=2):
        ws_g.cell(row=hdr, column=i, value=d).font = Font(bold=True, size=10)
    ws_g.cell(row=hdr, column=len(dates) + 2, value='合计').font = Font(bold=True)
    lines_seen = sorted({blk['line'] for blk in result['schedule']})
    for i, ln in enumerate(lines_seen, start=hdr + 1):
        ws_g.cell(row=i, column=1, value=ln)
        for j in range(2, len(dates) + 2):
            col = get_column_letter(j)
            ws_g.cell(row=i, column=j, value='=SUMIFS(排产明细!$L$2:$L{0},排产明细!$A$2:$A{0},$A{1},排产明细!$I$2:$I{0},{2}{3})'.format(last, i, col, hdr))
        ws_g.cell(row=i, column=len(dates) + 2, value='=SUM(B{0}:{1}{0})'.format(i, get_column_letter(len(dates) + 1)))
    rng = 'B{0}:{1}{2}'.format(hdr + 1, get_column_letter(len(dates) + 1), hdr + len(lines_seen))
    ws_g.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['8'], fill=PatternFill('solid', fgColor='FFC7CE')))
    ws_g.conditional_formatting.add(rng, CellIsRule(operator='between', formula=['4', '8'], fill=PatternFill('solid', fgColor='FFEB9C')))
    ws_g.column_dimensions['A'].width = 20
    for i in range(2, len(dates) + 3):
        ws_g.column_dimensions[get_column_letter(i)].width = 12
    ws_g.freeze_panes = 'B4'
    # 假设与口径
    ws_a = wb.create_sheet('假设与口径')
    ws_a.cell(row=1, column=1, value='排产假设与口径（结论可靠性 = 假设可核查）').font = tfont
    hdr2 = 4
    for c, h in enumerate(['假设ID', '项目', '取值', '来源', '校准状态'], start=1):
        cell = ws_a.cell(row=hdr2, column=c, value=h)
        cell.fill, cell.font, cell.border = hfill, hfont, border
    hyps = assumptions or [{'id': '-', 'item': '无额外假设', 'value': '-', 'source': '-', 'status': '-'}]
    for i, h in enumerate(hyps, start=hdr2 + 1):
        ws_a.cell(row=i, column=1, value=h.get('id', ''))
        ws_a.cell(row=i, column=2, value=h.get('item', ''))
        ws_a.cell(row=i, column=3, value=h.get('value', ''))
        ws_a.cell(row=i, column=4, value=h.get('source', ''))
        ws_a.cell(row=i, column=5, value=h.get('status', ''))
        for c in range(1, 6):
            ws_a.cell(row=i, column=c).border = border
    if sensitivity:
        sr = hdr2 + len(hyps) + 2
        ws_a.cell(row=sr, column=1, value='敏感度（准时率对假设的响应）').font = Font(bold=True)
        for k, v in sensitivity.items():
            sr += 1
            ws_a.cell(row=sr, column=1, value=k)
            ws_a.cell(row=sr, column=2, value=v)
    for col in 'ABCDE':
        ws_a.column_dimensions[col].width = 22
    wb.save(path)
    return path, checks
