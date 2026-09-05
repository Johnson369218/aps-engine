#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
参考试点 · 车间排产调度核心（qpf-schedule 技能的计算引擎）
=========================================================
引擎（自动选择）：
  cp        - OR-Tools CP-SAT（质量最高，需 ortools）
  heuristic - 纯 Python 启发式（EDD + 负载均衡 + 2-opt 局部搜索，零依赖）
两阶段设计：
  阶段一（排序/指派）：CP-SAT 用「可选区间 + 每线 NoOverlap」求出订单到产线的
        指派与先后顺序，目标 = 最小化加权拖期（优先级 1 最紧急，权重最大）。
  阶段二（精确化）：按真实换型时间矩阵（CIP/清洗/换型）重算每线任务墙钟时间，
        再做 2-opt 邻域搜索，在“不增加拖期”前提下尽量压缩换型时间。

输入（JSON）：
  lines.json    产线:  id/name/work_minutes_per_day/shift_start/first_date/weekends_off
  products.json 产品:  id/name/speed_per_hour(默认)/overhead_min/setup_min(换型矩阵)
  orders.json   订单:  id/product/qty/due/priority/allowed_lines(可选)/release(可选)

输出：
  schedule.json  每线任务序列（墙钟时间）、延期统计、瓶颈、产能利用率
用法：
  python3 scheduler.py --orders data/orders.json --lines data/lines.json \
      --products data/products.json --out output/schedule.json [--engine auto|cp|heuristic]
"""

import argparse
import json
import math
import random
import sys
from collections import defaultdict
from datetime import date, datetime, time, timedelta

try:
    from ortools.sat.python import cp_model
    HAVE_ORTools = True
except ImportError:  # pragma: no cover - 无 ortools 环境
    cp_model = None
    HAVE_ORTools = False

DEFAULT_SETUP_MIN = 30      # 未配置换型矩阵时的默认换型时间（分钟）
DEFAULT_OVERHEAD_MIN = 10   # 每批次的固定准备/收尾时间（分钟）


# ---------------------------------------------------------------- 工具 ----

def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def ensure(cond, msg):
    if not cond:
        raise ValueError(msg)


def parse_dt(s):
    """'YYYY-MM-DD' 或 'YYYY-MM-DD HH:MM' -> datetime"""
    s = s.strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"无法解析时间: {s}")


class Calendar:
    """产线日历：把「工作分钟」换算成墙钟时间；支持班次时长、周末休息、首日。"""

    def __init__(self, first_date, shift_start="08:00",
                 work_minutes_per_day=480, weekends_off=True):
        self.first = parse_dt(first_date).date()
        sh, sm = (int(x) for x in shift_start.split(":"))
        self.shift_start_min = sh * 60 + sm
        self.wmpd = int(work_minutes_per_day)
        self.weekends_off = bool(weekends_off)

    def _workday(self, d):
        return not (self.weekends_off and d.weekday() >= 5)

    def date_to_min(self, d):
        """从首日 00:00 到 d 日 00:00 的工作分钟数（天边界对齐）。"""
        off, cur = 0, self.first
        while cur < d:
            if self._workday(cur):
                off += self.wmpd
            cur += timedelta(days=1)
        return off

    def dt_to_min(self, dt):
        """墙钟时间 -> 工作分钟（用于交期/释放日换算）。"""
        base = self.date_to_min(dt.date())
        mofday = dt.hour * 60 + dt.minute
        return base + max(0, mofday - self.shift_start_min)

    def min_to_dt(self, m):
        """工作分钟 -> 墙钟 datetime。"""
        day, rem = divmod(int(m), self.wmpd)
        cur, skipped = self.first, 0
        while skipped < day:
            cur += timedelta(days=1)
            if self._workday(cur):
                skipped += 1
        return datetime.combine(cur, time(0, 0)) + timedelta(
            minutes=self.shift_start_min + rem)

    def fmt(self, m):
        return self.min_to_dt(m).strftime("%Y-%m-%d %H:%M")


# ---------------------------------------------------------------- 建模 ----

def get_setup(products, p_from, p_to):
    """换型时间（分钟）。同产品为 0；按矩阵查；缺省用默认值。"""
    if p_from == p_to:
        return 0
    m = products.get(p_from, {}).get("setup_min", {})
    if isinstance(m, dict):
        if p_to in m:
            return int(m[p_to])
        if p_from in products.get(p_to, {}).get("setup_min", {}):
            return int(products[p_to]["setup_min"][p_from])
    return int(products.get(p_from, {}).get("default_setup_min", DEFAULT_SETUP_MIN))


def get_speed(products, prod_id, line_id):
    prod = products.get(prod_id, {})
    speeds = prod.get("speeds", {})
    if isinstance(speeds, dict) and line_id in speeds:
        return float(speeds[line_id])
    return float(prod.get("speed_per_hour", 1000))


def build_durations(orders, products, allowed):
    """每订单在每可用产线上的加工分钟数。
    订单自带 duration_min 时直接采用（真实测时/折算时长）；否则按速度估算。"""
    durations = {}
    for o in orders:
        durations[o["id"]] = {}
        if o.get("duration_min"):
            for l in allowed[o["id"]]:
                durations[o["id"]][l] = int(o["duration_min"])
            continue
        for l in allowed[o["id"]]:
            prod = products.get(o["product"], {})
            cap8h = prod.get("capacity_8h") or prod.get("back_cap_8h")
            if cap8h:
                # 后工序（人力包装）决定整线入库：duration = qty × 480 / 8h产能
                proc = max(10, math.ceil(o["qty"] * 480 / cap8h))
            else:
                speed = get_speed(products, o["product"], l)
                proc = math.ceil(o["qty"] / speed * 60)
            proc += int(prod.get("overhead_min", 0))
            durations[o["id"]][l] = proc
    return durations


def build_front_durations(orders, products, allowed):
    """前工序（自动化设备加工）时长：duration = qty × 480 / front_cap_8h。
    前工序产能 ≥ 后工序（设备快），因此时长一般更短；前工序可提前做半成品。"""
    front = {}
    for o in orders:
        front[o["id"]] = {}
        for l in allowed[o["id"]]:
            prod = products.get(o["product"], {})
            fcap = prod.get("front_cap_8h") or prod.get("capacity_8h")
            if not fcap:
                # 无 8h 产能字段时用 speed_per_hour×8（每小时产能 → 8h 产能），避免兜底 1000 造成 80h 前工序
                fcap = (prod.get("speed_per_hour") or 1000) * 8
            front[o["id"]][l] = max(10, math.ceil(o["qty"] * 480 / fcap))
    return front


def compute_horizon(orders, durations, due_min, cal):
    """工作分钟上界：所有加工+换型之和的两倍 + 一个工作周缓冲。"""
    total = sum(durations[o["id"]][l]
                for o in orders for l in durations[o["id"]])
    return max(due_min.values()) + total * 2 + 5 * cal.wmpd


# ------------------------------------------------------------ 阶段一 ----

def solve_cp(orders, lines, allowed, durations, due_min, weights, cal,
             time_limit=20, cap_of=None, initial=None, seed=42):
    """CP-SAT：指派 + 排序，目标 = 最小化加权拖期。
    cap_of: 产线并行工位数（>1 用 cumulative 多工位并行，=1 用 NoOverlap 单机）。"""
    cap_of = cap_of or {}
    model = cp_model.CpModel()
    horizon = compute_horizon(orders, durations, due_min, cal)

    start, end, pres, iv = {}, {}, {}, {}
    release = {o["id"]: cal.dt_to_min(parse_dt(o.get("release", cal.first.isoformat())))
               for o in orders}

    if initial:
        # 热启动（Johnson/启发式初始解）：hint 只建议不强制，加速并稳定求解
        for o in orders:
            oid = o["id"]
            if oid in initial and initial[oid][0] in allowed[oid]:
                hint_l, hint_s, hint_e = initial[oid]
                pass  # 占位：hint 在变量创建后添加

    for o in orders:
        oid = o["id"]
        start[oid], end[oid], pres[oid], iv[oid] = {}, {}, {}, {}
        for l in allowed[oid]:
            dur = durations[oid][l]
            s = model.NewIntVar(0, horizon, f"s_{oid}_{l}")
            e = model.NewIntVar(0, horizon, f"e_{oid}_{l}")
            p = model.NewBoolVar(f"p_{oid}_{l}")
            model.Add(s >= release[oid]).OnlyEnforceIf(p)
            iv[oid][l] = model.NewOptionalIntervalVar(s, dur, e, p,
                                                      f"iv_{oid}_{l}")
            start[oid][l], end[oid][l], pres[oid][l] = s, e, p
        model.AddExactlyOne(pres[oid][l] for l in allowed[oid])
        if initial and oid in initial and initial[oid][0] in allowed[oid]:
            hint_l, hint_s, hint_e = initial[oid]
            model.AddHint(pres[oid][hint_l], 1)
            model.AddHint(start[oid][hint_l], max(0, min(int(hint_s), horizon)))
            model.AddHint(end[oid][hint_l], max(0, min(int(hint_e), horizon)))

    for l in lines:
        ivs = [iv[o["id"]][l] for o in orders if l in allowed[o["id"]]]
        cap = cap_of.get(l, 1)
        if cap <= 1:
            model.AddNoOverlap(ivs)
        else:
            # 多工位并行产线：同一时刻最多 cap 个任务并行（每人一个工位）
            model.AddCumulative(ivs, [1] * len(ivs), cap)

    # 加权拖期：tard >= end - due 仅在“订单被指派到该线”时生效（唯一成立）
    tard = {}
    for o in orders:
        oid = o["id"]
        t = model.NewIntVar(0, horizon, f"tard_{oid}")
        for l in allowed[oid]:
            model.Add(t >= end[oid][l] - due_min[oid]).OnlyEnforceIf(pres[oid][l])
        tard[oid] = t

    # 提前量惩罚（JIT/保质期约束）：early = due - end，
    # 目标里以远小于拖期的权重计入——避免产能富余时把任务全部提前堆到最早日，
    # 保证排产贴合交期分布（食品保质期红线：馒头/净菜不可长期提前生产）。
    EARLY_W = 0.05   # 提前 1 分钟 = 0.05×权重；拖期 1 分钟 = 1.0×权重（提前 20 分钟 ≈ 拖期 1 分钟）
    early = {}
    for o in orders:
        oid = o["id"]
        e = model.NewIntVar(0, horizon, f"early_{oid}")
        for l in allowed[oid]:
            model.Add(e >= due_min[oid] - end[oid][l]).OnlyEnforceIf(pres[oid][l])
        early[oid] = e

    obj = (sum(weights[oid] * tard[oid] for oid in tard)
           + EARLY_W * sum(weights[oid] * early[oid] for oid in early))
    model.Minimize(obj)

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit
    solver.parameters.random_seed = seed
    # B1 可复现：多 worker 并行即使固定 seed 也不确定；单 worker + 固定 seed 逐字节一致
    solver.parameters.num_search_workers = 1
    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return None, f"CP-SAT 无可行解 (status={solver.StatusName(status)})"

    seq = {l: [] for l in lines}
    for o in orders:
        oid = o["id"]
        chosen = next(l for l in allowed[oid]
                      if solver.Value(pres[oid][l]) == 1)
        seq[chosen].append({
            "order": oid, "product": o["product"], "qty": o["qty"],
            "due": o["due"], "priority": o["priority"],
            "proc_min": durations[oid][chosen],
            "start_min": solver.Value(start[oid][chosen]),
            "end_min": solver.Value(end[oid][chosen]),
            "release": o.get("release", cal.first.isoformat()),
            "order_time": o.get("order_time", o.get("due")),
        })
    for l in seq:
        seq[l].sort(key=lambda t: t["start_min"])
    return seq, None


def solve_heuristic(orders, lines, allowed, durations, due_min, weights, cal,
                    cap_of=None, seed=42):
    """启发式：优先级+EDD 排序 → 负载均衡指派 → 线内按换型最优插入。
    cap_of>1 时按多工位派发（任务放到最早空闲工位）。
    seed：排序/并列打破用 random.Random(seed)（当前确定性排序，seed 为兼容占位）。"""
    rng = random.Random(seed)  # noqa: F841  — 保留确定性随机源，供并列打破逻辑使用
    cap_of = cap_of or {}
    seq = {l: [] for l in lines}
    load = {l: 0 for l in lines}
    last_product = {l: None for l in lines}
    stations = {l: [0] * max(1, cap_of.get(l, 1)) for l in lines}  # 各工位可用时刻

    order_list = sorted(orders,
                        key=lambda o: (due_min[o["id"]], o["priority"],
                                       -o["qty"]))
    for o in order_list:
        oid = o["id"]
        # 选线：完成时间 + 换型 最靠前的线（权重 = 优先级）
        best_l, best_val = None, None
        for l in allowed[oid]:
            setup = get_setup_for_line(o["product"], last_product[l])
            val = load[l] + durations[oid][l] + setup
            if best_l is None or val < best_val:
                best_l, best_val = l, val
        dur = durations[oid][best_l]
        if len(stations[best_l]) > 1:
            # 多工位：最早空闲工位
            st = min(range(len(stations[best_l])),
                     key=lambda i: stations[best_l][i])
            start_min = stations[best_l][st]
            stations[best_l][st] = start_min + dur
            seq[best_l].append({
                "order": oid, "product": o["product"], "qty": o["qty"],
                "due": o["due"], "priority": o["priority"], "proc_min": dur,
                "start_min": start_min, "end_min": start_min + dur,
            })
        else:
            # 单机：线内插入，找换型增量最小的位置
            s = seq[best_l]
            setup_inc = []
            for i in range(len(s) + 1):
                before = s[i - 1]["product"] if i > 0 else last_product[best_l]
                after = s[i]["product"] if i < len(s) else None
                setup_inc.append(setup_of(before, o["product"], after))
            pos = min(range(len(setup_inc)), key=lambda i: setup_inc[i])
            start_min = (s[pos - 1]["end_min"] + setup_inc[pos]
                         if pos > 0 else stations[best_l][0])
            s.insert(pos, {"order": oid, "product": o["product"],
                           "qty": o["qty"], "due": o["due"],
                           "priority": o["priority"], "proc_min": dur,
                           "start_min": start_min, "end_min": start_min + dur})
            stations[best_l][0] = start_min + dur
        load[best_l] += dur
        last_product[best_l] = o["product"]
    for l in seq:
        seq[l].sort(key=lambda t: t["start_min"])
    return seq, None


def setup_of(before_prod, cur_prod, after_prod=None):
    """插入带来的换型增量：before→cur + cur→after - before→after。"""
    b, c, a = before_prod, cur_prod, after_prod
    if b is None:
        return 0 if a is None else 0  # 空线插入，无换型
    inc = get_setup_global(b, c)
    if a is not None:
        inc += get_setup_global(c, a) - get_setup_global(b, a)
    return max(0, inc)


# 全局换型查询（阶段一使用，避免传参混乱）
_G_PRODUCTS = {}


def _set_global_products(products):
    global _G_PRODUCTS
    _G_PRODUCTS = products


def get_setup_global(p_from, p_to):
    return get_setup(_G_PRODUCTS, p_from, p_to)


def get_setup_for_line(prod, last_prod):
    return 0 if last_prod is None else get_setup(_G_PRODUCTS, last_prod, prod)


# ------------------------------------------------------------ 阶段二 ----

def _proc_min_of(t):
    """任务加工分钟：优先用 backfill 携带的真实 proc_min；
    否则用 end-start 墙钟 - 换型（跨天任务墙钟跨度会虚高，仅兜底）。"""
    if t.get("proc_min") is not None:
        return int(t["proc_min"])
    s = parse_dt(t["start"]); e = parse_dt(t["end"])
    return max(0, int((e - s).total_seconds() / 60) - int(t.get("setup_min") or 0))


def backfill_front_ops(tasks, products, cal, back_start_of, front_durs, line_id):
    """前工序排产（食品安全合规）：
    - 前工序（自动化设备加工）可提前做半成品——半成品暂存合法；
    - 每个订单前工序必须在【后工序开始】前完成（截止=后工序开始）；
    - 前工序尽量晚做（贴后工序，缩短半成品存放），产能不足时向前溢出；
    - 前工序早于后工序的日期不视为“提前生产”（未入库，不涉及生产日期）。
    tasks: 订单列表（含 order/product）；back_start_of: order->后工序开始datetime；
    front_durs: order->(proc_front)；line_id: 产线 id。
    返回 {order: {"start": datetime, "end": datetime}}。
    """
    from collections import defaultdict
    wmpd = int(cal.wmpd)
    day_used = defaultdict(int)
    day_last_prod = {}
    out = {}
    for t in sorted(tasks, key=lambda x: (back_start_of[x["order"]], x.get("priority", 2))):
        oid = t["order"]
        proc = int(front_durs[oid])
        deadline = back_start_of[oid]
        # 从截止往前找（尽量晚，贴后工序；可提前溢出）
        placed = None
        d = deadline.date()
        while d >= cal.first:
            if cal._workday(d):
                prev = day_last_prod.get(d)
                setup = 0 if (prev is None or prev == t["product"]) \
                    else get_setup(products, prev, t["product"])
                if day_used[d] + proc + setup <= wmpd:
                    start_dt = (datetime.combine(d, time(8, 0))
                                + timedelta(minutes=day_used[d] + setup))
                    end_dt = start_dt + timedelta(minutes=proc)
                    if end_dt <= deadline:
                        day_used[d] += proc + setup
                        day_last_prod[d] = t["product"]
                        placed = (start_dt, end_dt)
                        break
            d -= timedelta(days=1)
        if placed is None:
            # 前工序极紧时允许顺延（一般不会发生，设备产能大）
            start_dt = deadline - timedelta(minutes=proc)
            placed = (start_dt, deadline)
        out[oid] = {"start": placed[0], "end": placed[1]}
    return out


def backfill_single_line(tasks, products, cal, line_names=None):
    """单线倒推填充（食品安全合规版）：
    - 每个订单优先在【交期当天】生产（生产日期=真实，禁止提前生产）；
    - 当天该线产能不足（480min 排满）时，才向前一天溢出（产能受限提前备货），
      并标注 early_days（提前天数）；
    - 任务支持跨天生产（时长 > 单日产能时自然分多日，如周订单）；
    - 提前到产线首日仍排不下 → 顺延到交期之后（延期，真实瓶颈）。
    tasks: 按交期升序排序的订单列表（含 order/product/qty/proc_min/due/priority）。
    """
    from collections import defaultdict
    wmpd = int(cal.wmpd)
    day_used = defaultdict(int)     # date -> 已用分钟（含换型）
    day_last_prod = {}              # date -> 当日最后一个产品（算换型）
    out, tardy = [], []

    def place(t, start_date):
        """从 start_date（含）开始放置任务，支持跨天。返回 (start_dt, end_dt, setup)。"""
        proc = int(t["proc_min"])
        d = start_date
        while not cal._workday(d):          # 找到第一个工作日
            d += timedelta(days=1)
        prev = day_last_prod.get(d)
        setup = (0 if (prev is None or prev == t["product"])
                 else get_setup(products, prev, t["product"]))
        start_dt = None
        remaining = proc
        while remaining > 0:
            if cal._workday(d):
                if start_dt is None:
                    if day_used[d] + setup > wmpd:   # 首日连换型都放不下 → 次日
                        d += timedelta(days=1)
                        continue
                    start_dt = (datetime.combine(d, time(8, 0))
                                + timedelta(minutes=day_used[d] + setup))
                    day_used[d] += setup
                    day_last_prod[d] = t["product"]
                avail = wmpd - day_used[d]
                use = min(avail, remaining)
                day_used[d] += use
                remaining -= use
                if remaining > 0:
                    d += timedelta(days=1)
            else:
                d += timedelta(days=1)
            if (d - start_date).days > 90:   # 保护：产能严重不足时避免死循环
                raise RuntimeError(f"backfill 溢出: 订单 {t['order']} 90天内排不下")
        # 最后一天的结束时刻
        last_day = d
        end_dt = (datetime.combine(last_day, time(8, 0))
                  + timedelta(minutes=day_used[last_day]))
        return start_dt, end_dt, setup

    for t in sorted(tasks, key=lambda x: (x["due"], x.get("priority", 2))):
        due_date = parse_dt(t["due"]).date()
        # ① 交期当天优先，向前溢出（产能受限提前）
        placed = None
        d = due_date
        while d >= cal.first:
            if cal._workday(d):
                # 试放：从 d 开始，看能否在 due 前完成（跨天允许）
                snap_used, snap_last = dict(day_used), dict(day_last_prod)
                start_dt, end_dt, setup = place(t, d)
                if end_dt <= parse_dt(t["due"]):
                    placed = (start_dt, end_dt, setup, d)
                    break
                # 试放失败 → 回滚占用
                day_used.clear(); day_used.update(snap_used)
                day_last_prod.clear(); day_last_prod.update(snap_last)
            d -= timedelta(days=1)
        if placed is None:
            # ② 顺延到交期后
            start_dt, end_dt, setup = place(t, due_date)
            placed = (start_dt, end_dt, setup, None)
        start_dt, end_dt, setup, used_from = placed
        # 入库时间 = 后工序完成时间（end_dt）；提前量按入库日 vs 交期日
        early_days = max(0, (due_date - end_dt.date()).days)
        tard_min = max(0, int((end_dt - parse_dt(t["due"])).total_seconds() / 60))
        rec = {
            "order": t["order"], "product": t["product"],
            "product_name": products.get(t["product"], {}).get("name", t["product"]),
            "qty": t["qty"], "proc_min": int(t["proc_min"]), "setup_min": setup,
            "setup_from": "-",
            "start": start_dt.strftime("%Y-%m-%d %H:%M"),
            "end": end_dt.strftime("%Y-%m-%d %H:%M"),
            "prod_date": end_dt.strftime("%Y-%m-%d"),
            "due": t["due"], "tardy_min": tard_min, "early_days": early_days,
            "order_time": t.get("order_time", t["due"]),
        }
        ot = t.get("order_time")
        if ot and str(ot).strip():
            try:
                rec["order_lead_h"] = round((end_dt - parse_dt(str(ot))).total_seconds() / 3600, 1)
            except Exception:
                rec["order_lead_h"] = None
        else:
            rec["order_lead_h"] = None   # 无 ERP 下单时间（如 6/7 月/预测订单）不显示时差
        out.append(rec)
        if tard_min > 0:
            tardy.append({"order": t["order"], "product": t["product"],
                          "product_name": rec["product_name"],
                          "tardy_min": tard_min, "due": t["due"],
                          "end": rec["end"]})
    return out, tardy


def finalize(seq, products, cal, line_names=None, cap_of=None, front_durs=None):
    """把每线任务序列转成墙钟排产表。
    - 单机模式（全部线 capacity<=1）：按真实换型矩阵串行重算墙钟。
    - 多工位模式（任一线 capacity>1）：直接用引擎给出的 start_min/end_min
      （并行工位由引擎保证），换型在工位间消化、不硬建模。
    """
    line_names = line_names or {}
    cap_of = cap_of or {}
    multi = any(cap_of.get(l, 1) > 1 for l in seq)
    schedule, total_setup = [], 0
    tardy, util = [], {}
    for l in seq:
        tasks = seq[l]
        if not multi:
            # 单机串行重算：按交期升序（EDD）处理，保证早期订单不被晚期订单挤占
            tasks = sorted(tasks, key=lambda t: (t["due"], t.get("priority", 2)))
        proc_total = 0
        if not multi:
            # 食品安全合规：倒推填充（当天生产优先，产能受限才提前）——后工序
            line_tasks, line_tardy = backfill_single_line(tasks, products, cal)
            tardy.extend(line_tardy)
            total_setup += sum(int(x["setup_min"]) for x in line_tasks)
            proc_total = sum(_proc_min_of(x) for x in line_tasks)
            # 前工序（自动化设备加工）排产：截止=后工序开始，可提前做半成品
            if front_durs:
                back_start_of = {x["order"]: parse_dt(x["start"]) for x in line_tasks}
                fr = backfill_front_ops(tasks, products, cal, back_start_of,
                                        {oid: front_durs.get(oid, {}).get(l, 10)
                                         for oid in front_durs}, l)
                for x in line_tasks:
                    f = fr.get(x["order"])
                    if f:
                        x["front_start"] = f["start"].strftime("%Y-%m-%d %H:%M")
                        x["front_end"] = f["end"].strftime("%Y-%m-%d %H:%M")
        else:
            prev_prod, t_end = None, 0
            line_tasks = []
            for t in tasks:
                setup, s0, e0 = 0, t["start_min"], t["end_min"]
                setup_from = "-"
                total_setup += setup
                # 延期按墙钟分钟（与单机 backfill 口径一致；周末不计入工作分钟）
                tard_min = max(0, int((parse_dt(cal.fmt(e0))
                                       - parse_dt(t["due"])).total_seconds() / 60))
                end_str = cal.fmt(e0)
                due_date = parse_dt(t["due"]).date()
                early = max(0, (due_date - parse_dt(end_str).date()).days)
                line_tasks.append({
                    "order": t["order"], "product": t["product"],
                    "product_name": products.get(t["product"], {}).get("name", t["product"]),
                    "qty": t["qty"], "setup_min": setup, "setup_from": setup_from,
                    "start": cal.fmt(s0), "end": end_str, "prod_date": end_str[:10],
                    "due": t["due"], "tardy_min": tard_min, "early_days": early,
                    "front_start": None, "front_end": None,
                    "order_time": t.get("order_time"),
                    "order_lead_h": None,
                })
                if tard_min > 0:
                    tardy.append({"order": t["order"], "product": t["product"],
                                  "product_name": line_tasks[-1]["product_name"],
                                  "tardy_min": tard_min, "due": t["due"],
                                  "end": line_tasks[-1]["end"]})
                prev_prod = t["product"]
                t_end = e0
        # 利用率：以“覆盖的首日~末日”为窗口估算（多工位按 容量×窗口 折算负荷）
        if line_tasks:
            # 任务顺序≠时间顺序（backfill 输出按放置次序）→ 用 min/max 求覆盖窗口
            d0 = min(parse_dt(x["start"]).date() for x in line_tasks)
            d1 = max(parse_dt(x["end"]).date() for x in line_tasks)
            days = cal.date_to_min(d1) - cal.date_to_min(d0) + cal.wmpd
            cap = max(1, cap_of.get(l, 1))
            util[l] = round(proc_total / max(days * cap, 1), 3)
        schedule.append({"line": l, "line_name": line_names.get(l, l),
                         "tasks": line_tasks})
    return schedule, tardy, total_setup, util


def local_improve(seq, products, due_min, weights, cal, passes=3):
    """2-opt 邻域搜索：交换同线相邻任务，若加权拖期不增且换型减少则接受。"""
    for _ in range(passes):
        improved = False
        for l, tasks in seq.items():
            for i in range(len(tasks) - 1):
                a, b = tasks[i], tasks[i + 1]
                if a["product"] == b["product"]:
                    continue
                # 试交换
                cand = tasks[:]
                cand[i], cand[i + 1] = cand[i + 1], cand[i]
                if _line_value(cand, products, due_min, weights, cal) < \
                        _line_value(tasks, products, due_min, weights, cal):
                    seq[l] = cand
                    improved = True
        if not improved:
            break
    return seq


def _line_value(tasks, products, due_min, weights, cal):
    """线内任务序列的评价值：加权拖期为主，换型分钟为辅。"""
    tv = wv = 0
    prev = None
    t = 0
    for x in tasks:
        setup = 0 if prev is None else get_setup(products, prev, x["product"])
        t += setup + x["proc_min"]
        due = due_min[x["order"]]
        wv += weights[x["order"]] * max(0, t - due)
        tv += setup
        prev = x["product"]
    return wv * 10000 + tv


# ---------------------------------------------------------------- 汇总 ----

def summarize(orders, schedule, tardy, total_setup, util):
    n = len(orders)
    total_tard = sum(x["tardy_min"] for x in tardy)
    max_tard = max((x["tardy_min"] for x in tardy), default=0)
    on_time = n - len(tardy)
    return {
        "orders": n,
        "on_time": on_time,
        "tardy": len(tardy),
        "on_time_rate": round(on_time / n, 3) if n else 1.0,
        "total_tardiness_min": total_tard,
        "max_tardiness_min": max_tard,
        "total_setup_min": total_setup,
        "utilization": util,
        "bottlenecks": sorted(
            [l for l, u in util.items() if u >= 0.8],
            key=lambda l: -util[l]),
    }


# ---------------------------------------------------------------- 主流程 ----

def run(orders, lines, products, engine="auto", time_limit=20, warm_start=True, seed=42):
    # 归一化：products 允许为列表（按 id 转字典），lines 保持列表
    if isinstance(products, list):
        products = {p["id"]: p for p in products}
    _set_global_products(products)
    line_ids = [l["id"] for l in lines]
    cal = {l["id"]: Calendar(l.get("first_date", date.today().isoformat()),
                             l.get("shift_start", "08:00"),
                             l.get("work_minutes_per_day", 480),
                             l.get("weekends_off", True))
           for l in lines}

    allowed = {}
    for o in orders:
        al = o.get("allowed_lines") or line_ids
        ensure(set(al) <= set(line_ids),
               f"订单 {o['id']} 的 allowed_lines 含未知产线: {al}")
        allowed[o["id"]] = al

    durations = build_durations(orders, products, allowed)
    # 交期/释放日按产线日历换算（各线日历一致时无差异）
    first_cal = cal[line_ids[0]]
    due_min = {o["id"]: first_cal.dt_to_min(parse_dt(o["due"]))
               for o in orders}
    weights = {o["id"]: {1: 10, 2: 4, 3: 1}.get(int(o.get("priority", 2)), 4)
               for o in orders}

    engine = engine.lower()
    if engine == "auto":
        engine = "cp" if HAVE_ORTools else "heuristic"

    cap_of = {l["id"]: int(l.get("capacity", 1)) for l in lines}
    multi = any(c > 1 for c in cap_of.values())

    if engine == "cp" and HAVE_ORTools:
        initial = None
        if warm_start:
            # 启发式初始解作为 CP-SAT 热启动（创新点1：Johnson/启发式 warm start）
            heu_seq, _ = solve_heuristic(orders, line_ids, allowed, durations,
                                         due_min, weights, first_cal, cap_of, seed)
            initial = {t["order"]: (l, t["start_min"], t["end_min"])
                       for l, ts in heu_seq.items() for t in ts}
        seq, err = solve_cp(orders, line_ids, allowed, durations, due_min,
                            weights, first_cal, time_limit, cap_of, initial, seed)
        if seq is None:
            sys.stderr.write(f"警告: CP-SAT 失败（{err}），回退启发式。\n")
            seq, _ = solve_heuristic(orders, line_ids, allowed, durations,
                                     due_min, weights, first_cal, cap_of, seed)
            engine = "heuristic"
    else:
        seq, _ = solve_heuristic(orders, line_ids, allowed, durations, due_min,
                                 weights, first_cal, cap_of, seed)

    line_names = {l["id"]: l["name"] for l in lines}
    if not multi:
        # 单机模式才做换型 2-opt 局部优化
        seq = local_improve(seq, products, due_min, weights, first_cal)
    front_durs = build_front_durations(orders, products, allowed)
    schedule, tardy, total_setup, util = finalize(seq, products, first_cal,
                                                  line_names, cap_of, front_durs)
    summary = summarize(orders, schedule, tardy, total_setup, util)

    result = {
        "engine": engine,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "summary": summary,
        "tardy_orders": sorted(tardy, key=lambda x: -x["tardy_min"]),
        "schedule": schedule,
    }
    return result


# ---------------------------------------------------------------- xlsx ----

def export_xlsx(result, path):
    """把排产结果导出为 .xlsx（排产计划 + 汇总两个 Sheet）。需要 openpyxl。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    _palette = ["FCE4D6", "E2EFDA", "DDEBF7", "FFF2CC", "E4DFEC", "D9E1F2"]
    line_fills = {f"L{i + 1}": PatternFill("solid", fgColor=c)
                  for i, c in enumerate(_palette)}
    tardy_fill = PatternFill("solid", fgColor="FFC7CE")
    tardy_font = Font(color="9C0006")

    # ---- Sheet1 排产计划 ----
    ws = wb.active
    ws.title = "排产计划"
    s = result["summary"]
    ws.merge_cells("A1:J1")
    ws["A1"] = "参考试点 2026-07 车间排产计划"
    ws["A1"].font = title_font
    ws.merge_cells("A2:J2")
    ws["A2"] = (f"引擎: {result['engine'].upper()} | 生成时间: {result['generated_at']} | "
                f"准时率 {s['on_time_rate']:.0%}（准时 {s['on_time']}/{s['orders']}）| "
                f"总延期 {s['total_tardiness_min']} 分钟 | 换型总耗时 {s['total_setup_min']} 分钟")
    has_origin = any(t.get("origin") for blk in result["schedule"] for t in blk["tasks"])
    headers = ["产线", "订单", "产品", "数量", "换型(分)", "上一产品",
               "生产日期(入库)", "前工序", "后工序-开始", "后工序-结束", "交期",
               "订单生成时间", "下单→入库(时)", "提前(天)", "延期(分)"]
    if has_origin:
        headers.append("类型")
    hr = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.fill, cell.font, cell.border = header_fill, header_font, border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r = hr + 1
    for blk in result["schedule"]:
        line_id = blk["line"]
        for t in blk["tasks"]:
            fs_ = t.get("front_start")
            front_cell = (fs_ + " ~ " + t.get("front_end", "")) if fs_ else "-"
            vals = [blk.get("line_name", line_id), t["order"],
                    t.get("product_name", t["product"]),
                    t["qty"], t["setup_min"], t.get("setup_from", ""),
                    t.get("prod_date", t["start"][:10]),
                    front_cell,
                    t["start"], t["end"], t["due"],
                    t.get("order_time", ""), t.get("order_lead_h", ""),
                    t.get("early_days", 0), t["tardy_min"]]
            if has_origin:
                vals.append(t.get("origin", ""))
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = line_fills.get(line_id)
                if t["tardy_min"] > 0 and c in (9, 12):
                    cell.fill, cell.font = tardy_fill, tardy_font
            r += 1
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = {
            1: 14, 2: 8, 3: 14, 4: 8, 5: 9, 6: 9, 7: 12, 8: 18, 9: 18, 10: 18,
            11: 9, 12: 9,
        }.get(c, 12)
    ws.freeze_panes = "A5"

    # ---- Sheet2 汇总 ----
    ws2 = wb.create_sheet("汇总")
    ws2.merge_cells("A1:B1")
    ws2["A1"] = "排产汇总"
    ws2["A1"].font = title_font
    rows = [
        ("订单总数", s["orders"]),
        ("准时订单", s["on_time"]),
        ("延期订单", s["tardy"]),
        ("准时率", f"{s['on_time_rate']:.1%}"),
        ("总延期(分钟)", s["total_tardiness_min"]),
        ("最大延期(分钟)", s["max_tardiness_min"]),
        ("换型总耗时(分钟)", s["total_setup_min"]),
        ("瓶颈/高负荷产线", "、".join(s["bottlenecks"]) or "无"),
    ]
    r = 3
    for k, v in rows:
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
        for c in (1, 2):
            ws2.cell(row=r, column=c).border = border
        r += 1
    r += 1
    ws2.cell(row=r, column=1, value="产线利用率").font = Font(bold=True)
    r += 1
    name_of = {blk["line"]: blk.get("line_name", blk["line"])
               for blk in result["schedule"]}
    for lid, u in s["utilization"].items():
        ws2.cell(row=r, column=1, value=name_of.get(lid, lid))
        ws2.cell(row=r, column=2, value=u).number_format = "0.0%"
        r += 1
    if result["tardy_orders"]:
        r += 1
        ws2.cell(row=r, column=1, value="延期订单明细").font = Font(bold=True)
        r += 1
        for t in result["tardy_orders"]:
            ws2.cell(row=r, column=1, value=t["order"])
            ws2.cell(row=r, column=2,
                     value=f"{t['product']} 延期 {t['tardy_min']} 分钟（交期 {t['due']}）")
            r += 1
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 44

    wb.save(path)


def main(argv=None):
    ap = argparse.ArgumentParser(description="参考试点车间排产调度核心")
    ap.add_argument("--orders", required=True)
    ap.add_argument("--lines", required=True)
    ap.add_argument("--products", required=True)
    ap.add_argument("--out", default="output/schedule.json")
    ap.add_argument("--xlsx", default=None, help="同时导出排产表 .xlsx")
    ap.add_argument("--engine", default="auto",
                    choices=["auto", "cp", "heuristic"])
    ap.add_argument("--time-limit", type=int, default=20)
    args = ap.parse_args(argv)

    result = run(load_json(args.orders), load_json(args.lines),
                 load_json(args.products),
                 engine=args.engine, time_limit=args.time_limit)

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    if args.xlsx:
        try:
            export_xlsx(result, args.xlsx)
        except ImportError:
            sys.stderr.write("提示: 未安装 openpyxl，跳过 xlsx 导出。"
                             "安装: python -m pip install openpyxl\n")

    s = result["summary"]
    print(f"引擎: {result['engine']} | 订单 {s['orders']} 个 | "
          f"准时 {s['on_time']} / 延期 {s['tardy']}（准时率 {s['on_time_rate']:.0%}）")
    print(f"总延期 {s['total_tardiness_min']} 分钟，最大延期 {s['max_tardiness_min']} 分钟，"
          f"换型总耗时 {s['total_setup_min']} 分钟")
    print("瓶颈/高负荷产线:", "、".join(s["bottlenecks"]) or "无")
    for blk in result["schedule"]:
        print(f"\n【{blk.get('line_name', blk['line'])}】利用率 "
              f"{s['utilization'].get(blk['line'], 0):.0%}")
        for t in blk["tasks"]:
            flag = " ⚠延期" if t["tardy_min"] > 0 else ""
            print(f"  {t['start']} → {t['end']}  {t.get('product_name', t['product'])}×{t['qty']} "
                  f"(换型 {t['setup_min']}min){flag}")
    print(f"\n已写入: {args.out}")


if __name__ == "__main__":
    main()
