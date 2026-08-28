# -*- coding: utf-8 -*-
"""排产表导出：21 列口径对齐《07_日计划/10_班前执行单》（调研报告主工作簿）。"""
import math
from datetime import datetime

COLUMNS = [
    "生产日期", "周编号", "月份", "批次序号", "品类", "SKU_ID", "产品名称",
    "工作中心", "单位", "本批计划数量", "计划产值", "8h计划产能", "负荷小时",
    "需并行班组", "标准班组人数", "建议直接人数", "可用并行班组", "风险状态",
    "前工序", "后工序-开始", "后工序-结束", "入库日期", "交期", "延期(分)",
    "达成率", "执行状态", "异常原因",
]


def _iso_week(d):
    try:
        return datetime.strptime(d, "%Y-%m-%d").isocalendar()[1]
    except ValueError:
        return ""


def export_schedule_xlsx(result, path, products=None):
    """导出 21+ 列排产表（Sheet1 排产计划 / Sheet2 汇总）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    prod_by_id = {}
    if isinstance(products, list):
        prod_by_id = {p["id"]: p for p in products}
    elif isinstance(products, dict):
        prod_by_id = products

    wb = Workbook()
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)

    ws = wb.active
    ws.title = "排产计划"
    s = result["summary"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.cell(row=1, column=1, value="岐品福 车间排产计划（对齐 07_日计划/班前执行单 口径）").font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    ws.cell(row=2, column=1, value=(
        f"引擎: {result['engine'].upper()} | {result['generated_at']} | "
        f"订单 {s['orders']} | 准时率 {s['on_time_rate']:.0%} | 总延期 {s['total_tardiness_min']}分 | "
        f"换型 {s['total_setup_min']}分 | 入库=后工序完成日（真实生产日期）"))

    hr = 3
    for c, h in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.fill, cell.font, cell.border = header_fill, header_font, border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def cap_8h(product):
        if not product:
            return None
        v = product.get("capacity_8h")
        if v:
            return float(v)
        spd = product.get("speed_per_hour")
        return float(spd) * 8 if spd else None

    def staff_of(product):
        return product.get("staff") if product else None

    r = hr + 1
    for line in result["schedule"]:
        for i, t in enumerate(line["tasks"], start=1):
            prod = prod_by_id.get(t.get("product"))
            c8 = cap_8h(prod)
            load = round(t["qty"] / c8 * 8, 2) if c8 else ""
            n_shift = math.ceil(load / 8) if isinstance(load, (int, float)) else ""
            staff = staff_of(prod)
            sugg = (staff * n_shift) if (staff and isinstance(n_shift, int)) else ""
            risk = "产能冲突" if isinstance(load, (int, float)) and load > 8 else ""
            price = prod.get("price") if prod else None
            val = [
                t.get("prod_date", ""), _iso_week(t.get("prod_date", "")),
                (t.get("prod_date") or "")[:7].replace("-", ""),
                i,
                (prod or {}).get("category", ""),
                t.get("product", ""), t.get("product_name", ""),
                line.get("line_name", line["line"]),
                (prod or {}).get("unit", ""),
                t.get("qty", ""),
                round(t["qty"] * price, 2) if price else "",
                c8 if c8 else "", load, n_shift, staff, sugg,
                (prod or {}).get("staff", ""), risk,
                f"{t.get('front_start','')}→{t.get('front_end','')}",
                t.get("start", ""), t.get("end", ""), t.get("prod_date", ""),
                t.get("due", ""), t.get("tardy_min", 0),
                "", "", "",
            ]
            for c, v in enumerate(val, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = border
            r += 1

    widths = [12, 8, 8, 8, 12, 10, 20, 12, 8, 14, 12, 12, 10, 10, 10, 12, 10, 10,
              24, 18, 18, 12, 18, 8, 8, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("汇总")
    ws2.append(["指标", "值"])
    ws2.append(["订单数", s["orders"]])
    ws2.append(["准时", s["on_time"]])
    ws2.append(["延期", s["tardy"]])
    ws2.append(["准时率", f"{s['on_time_rate']:.1%}"])
    ws2.append(["总延期(分)", s["total_tardiness_min"]])
    ws2.append(["换型总耗时(分)", s["total_setup_min"]])
    ws2.append(["瓶颈/高负荷(≥80%)", "、".join(s["bottlenecks"]) or "无"])
    ws2.append(["", ""])
    ws2.append(["产线", "利用率"])
    for k, v in s["utilization"].items():
        ws2.append([k, f"{v:.1%}"])
    for col in ("A", "B"):
        ws2.column_dimensions[col].width = 24

    wb.save(path)
    return path
