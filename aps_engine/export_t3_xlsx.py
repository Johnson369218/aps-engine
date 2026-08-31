# -*- coding: utf-8 -*-
"""T+3 排产表（Excel 公式联动版）生成器。

Sheet1 T+3排产计划：明细（27 列口径 + 提前天数 + 订单类型）
Sheet2 主数据：SKU 主数据（VLOOKUP 数据源）
Sheet3 汇总：KPI + 按日 + 按产线（SUMIFS/COUNTIFS 勾稽明细）
Sheet4 甘特图：工作中心×日期 负荷矩阵（复刻主工作簿 08_甘特图：单元格=计划负荷小时），
              每线两行（负荷小时 + 需并行班组），条件格式预警，底行合计
fullCalcOnLoad=True：Excel/WPS 打开即自动重算。
"""
import math
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DETAIL_COLS = [
    "生产日期", "星期", "周编号", "批次", "冻结状态", "SKU_ID", "产品名称",
    "工作中心", "单位", "计划数量", "单价", "计划产值", "8h计划产能", "负荷小时",
    "需并行班组", "标准班组人数", "建议直接人数", "风险状态", "前工序",
    "后工序-开始", "后工序-结束", "交期", "提前天数", "订单类型",
]

def _detail_formulas(r):
    return {
        "B": f"=TEXT(A{r},\"aaaa\")",
        "C": f"=ISOWEEKNUM(A{r})",
        "G": f"=VLOOKUP(F{r},主数据!$A:$F,2,0)",
        "I": f"=VLOOKUP(F{r},主数据!$A:$F,3,0)",
        "K": f"=VLOOKUP(F{r},主数据!$A:$F,4,0)",
        "L": f"=ROUND(J{r}*K{r},2)",
        "M": f"=VLOOKUP(F{r},主数据!$A:$F,5,0)",
        "N": f"=ROUND(J{r}/M{r}*8,2)",
        "O": f"=MAX(1,ROUNDUP(N{r}/8,0))",
        "P": f"=VLOOKUP(F{r},主数据!$A:$F,6,0)",
        "Q": f"=IF(P" + str(r) + "=\"\",\"\",P" + str(r) + "*O" + str(r) + ")",
        "R": f"=IF(N{r}>8,\"⚠产能冲突\",\"\")",
    }

def _rhup(x, nd=2):
    return float(Decimal(str(x)).quantize(Decimal("1." + "0" * nd), rounding=ROUND_HALF_UP))


def export_t3_xlsx(result, products, lines, dates, freeze_map, path):
    """生成 T+3 公式联动 Excel（明细/主数据/汇总/甘特图）。返回 (path, checks)。"""
    prod_by_id = {p["id"]: p for p in products}
    line_names = {l["id"]: l["name"] for l in lines}
    checks = []

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=13)

    # ── 主数据 ──
    ws_m = wb.create_sheet("主数据")
    for c, h in enumerate(["SKU_ID", "产品名称", "单位", "单价", "8h计划产能", "标准班组人数"], start=1):
        cell = ws_m.cell(row=1, column=c, value=h)
        cell.fill, cell.font, cell.border = header_fill, header_font, border
    for i, p in enumerate(sorted(products, key=lambda x: x["id"]), start=2):
        ws_m.cell(row=i, column=1, value=p["id"])
        ws_m.cell(row=i, column=2, value=p.get("name", ""))
        ws_m.cell(row=i, column=3, value=p.get("unit", ""))
        ws_m.cell(row=i, column=4, value=p.get("price", 0))
        ws_m.cell(row=i, column=5, value=p.get("capacity_8h", ""))
        ws_m.cell(row=i, column=6, value=p.get("staff") or "")
    for col in "ABCDEF":
        ws_m.column_dimensions[col].width = 16
    ws_m.freeze_panes = "A2"

    # ── 明细 ──
    ws = wb.active
    ws.title = "T+3排产计划"
    for c, h in enumerate(DETAIL_COLS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font, cell.border = header_fill, header_font, border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    r = 2
    line_seq = defaultdict(int)
    for blk in result["schedule"]:
        line = blk["line"]
        for t in blk["tasks"]:
            date = t["prod_date"]
            sku = t["product"]
            p = prod_by_id.get(sku, {})
            qty = float(t["qty"])
            price = float(p.get("price") or 0)
            cap8 = float(p.get("capacity_8h") or 0)
            exp_L = _rhup(qty * price)
            exp_N = _rhup(qty / cap8 * 8) if cap8 else 0
            exp_O = max(1, math.ceil(exp_N / 8))
            staff = p.get("staff") or ""
            exp_Q = (staff * exp_O) if staff else ""
            exp_R = "⚠产能冲突" if exp_N > 8 else ""
            line_seq[line] += 1
            vals = {
                "A": date,
                "D": line_seq[line],
                "E": freeze_map.get(date, "滚动"),
                "F": sku,
                "H": line_names.get(line, line),
                "J": t["qty"],
                "S": f"{t.get('front_start', '')}→{t.get('front_end', '')}",
                "T": t.get("start", ""),
                "U": t.get("end", ""),
                "V": t.get("due", ""),
                "W": t.get("early_days", 0),
                "X": t.get("order_type", t.get("origin", "confirmed")),
            }
            for col, v in vals.items():
                ws.cell(row=r, column=ord(col) - 64, value=v).border = border
            for col, formula in _detail_formulas(r).items():
                ws.cell(row=r, column=ord(col) - 64, value=formula).border = border
                if col in ("L", "N", "O", "Q", "R"):
                    expect = {"L": exp_L, "N": exp_N, "O": exp_O, "Q": exp_Q, "R": exp_R}[col]
                    checks.append((col, r, formula, expect))
            r += 1
    last = r - 1
    widths = [12, 10, 8, 8, 10, 10, 22, 14, 8, 12, 8, 12, 12, 10, 12, 12, 12, 12, 26, 18, 18, 18, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_COLS))}{last}"

    # ── 汇总（显式行号，避免覆盖）──
    ws_s = wb.create_sheet("汇总")
    ws_s.cell(row=1, column=1, value="T+3 排产汇总（全部公式联动明细）").font = title_font
    ws_s.cell(row=2, column=1, value=f"生成 {result['generated_at']} | 引擎 {result['engine']} | 窗口 {dates[0]} ~ {dates[-1]} | 提前区 8/16-8/27（产能受限提前）")
    ws_s.cell(row=3, column=1, value="指标").font = Font(bold=True)
    ws_s.cell(row=3, column=2, value="公式").font = Font(bold=True)
    ws_s.cell(row=3, column=3, value="口径").font = Font(bold=True)
    kpis = [
        ("总订单数", "=COUNTIF(T+3排产计划!$X:$X,\"forecast\")+COUNTIF(T+3排产计划!$X:$X,\"confirmed\")", "明细行数"),
        ("计划总量", "=SUM(T+3排产计划!$J:$J)", "混合单位合计"),
        ("总负荷(小时)", "=SUM(T+3排产计划!$N:$N)", "Σ(数量÷8h产能×8)"),
        ("总计划产值", "=SUM(T+3排产计划!$L:$L)", "Σ(数量×单价)"),
        ("提前生产单数", "=COUNTIF(T+3排产计划!$W:$W,\">0\")", "产能受限提前，见明细提前天数列"),
    ]
    for i, (k, fml, note) in enumerate(kpis, start=4):
        ws_s.cell(row=i, column=1, value=k)
        ws_s.cell(row=i, column=2, value=fml)
        ws_s.cell(row=i, column=3, value=note)
    # 按日
    row = 10
    ws_s.cell(row=row, column=1, value="日期").font = Font(bold=True)
    ws_s.cell(row=row, column=2, value="订单数").font = Font(bold=True)
    ws_s.cell(row=row, column=3, value="计划数量").font = Font(bold=True)
    ws_s.cell(row=row, column=4, value="负荷小时").font = Font(bold=True)
    ws_s.cell(row=row, column=5, value="产能(6线×8h)").font = Font(bold=True)
    ws_s.cell(row=row, column=6, value="利用率").font = Font(bold=True)
    for i, d in enumerate(dates, start=row + 1):
        ws_s.cell(row=i, column=1, value=d)
        ws_s.cell(row=i, column=2, value=f"=COUNTIFS(T+3排产计划!$A:$A,$A{i})")
        ws_s.cell(row=i, column=3, value=f"=SUMIFS(T+3排产计划!$J:$J,T+3排产计划!$A:$A,$A{i})")
        ws_s.cell(row=i, column=4, value=f"=SUMIFS(T+3排产计划!$N:$N,T+3排产计划!$A:$A,$A{i})")
        ws_s.cell(row=i, column=5, value=48)
        ws_s.cell(row=i, column=6, value=f"=IFERROR(ROUND(D{i}/E{i},3),0)")
    # 按产线
    row2 = row + 1 + len(dates) + 1
    ws_s.cell(row=row2, column=1, value="产线").font = Font(bold=True)
    ws_s.cell(row=row2, column=2, value="任务数").font = Font(bold=True)
    ws_s.cell(row=row2, column=3, value="负荷小时").font = Font(bold=True)
    ws_s.cell(row=row2, column=4, value="利用率(3天×8h)").font = Font(bold=True)
    ws_s.cell(row=row2, column=5, value="判断").font = Font(bold=True)
    for i, ln in enumerate(sorted(line_names), start=row2 + 1):
        ws_s.cell(row=i, column=1, value=line_names[ln])
        ws_s.cell(row=i, column=2, value=f"=COUNTIFS(T+3排产计划!$H:$H,$A{i})")
        ws_s.cell(row=i, column=3, value=f"=SUMIFS(T+3排产计划!$N:$N,T+3排产计划!$H:$H,$A{i})")
        ws_s.cell(row=i, column=4, value=f"=IFERROR(ROUND(C{i}/(8*{len(dates)}),3),0)")
        ws_s.cell(row=i, column=5, value=f"=IF(D" + str(i) + ">0.8,\"⚠ 高负荷\",IF(D" + str(i) + ">0.5,\"◐ 中负荷\",\"○ 低负荷\"))")
    for col in "ABCDEF":
        ws_s.column_dimensions[col].width = 18

    # ── 甘特图（工作中心 × 日期 负荷矩阵，复刻主工作簿 08_甘特图）──
    ws_g = wb.create_sheet("甘特图")
    dates_all = sorted({t["prod_date"] for blk in result["schedule"] for t in blk["tasks"]})
    ws_g.cell(row=1, column=1, value="工作中心 × 日期 负荷矩阵（单元格 = 计划负荷小时；单线单班上限 8h）").font = title_font
    ws_g.cell(row=2, column=1, value="需并行班组 = MAX(1,ROUNDUP(负荷/8,0))；红>8h（超单班） 黄4-8h 绿<4h 空白=无计划").font = Font(size=10, color="666666")
    hdr = 3
    ws_g.cell(row=hdr, column=1, value="工作中心 / 日期").font = Font(bold=True)
    for i, d in enumerate(dates_all, start=2):
        cell = ws_g.cell(row=hdr, column=i, value=d)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
    ws_g.cell(row=hdr, column=len(dates_all) + 2, value="合计").font = Font(bold=True)
    r = hdr + 1
    for ln in sorted(line_names):
        name = line_names[ln]
        # 负荷行
        ws_g.cell(row=r, column=1, value=f"{ln} {name}")
        for i, d in enumerate(dates_all, start=2):
            cell = ws_g.cell(row=r, column=i,
                value=f"=SUMIFS(T+3排产计划!$N:$N,T+3排产计划!$A:$A,$B${hdr},T+3排产计划!$H:$H,$A{r})")
            cell.alignment = Alignment(horizontal="center")
        ws_g.cell(row=r, column=len(dates_all) + 2, value=f"=SUM(B{r}:{get_column_letter(len(dates_all) + 1)}{r})")
        r += 1
        # 并行班组行
        ws_g.cell(row=r, column=1, value=f"{ln} 需并行班组").font = Font(size=10, color="666666")
        for i, d in enumerate(dates_all, start=2):
            cell = ws_g.cell(row=r, column=i,
                value=f"=IF(B" + str(r - 1) + "=\"\",\"\",MAX(1,ROUNDUP(B" + str(r - 1) + "/8,0)))")
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=10, color="666666")
        r += 1
    # 合计行：显式累加各线负荷行（hdr+1 起，步长 2）
    load_rows = list(range(hdr + 1, r - 1, 2))
    ws_g.cell(row=r, column=1, value="当日合计负荷").font = Font(bold=True)
    for i in range(2, len(dates_all) + 2):
        col = get_column_letter(i)
        refs = "+".join(f"{col}{x}" for x in load_rows)
        cell = ws_g.cell(row=r, column=i, value=f"={refs}")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    # 条件格式：红>8 黄4-8 绿0-4
    rng = f"B{hdr + 1}:{get_column_letter(len(dates_all) + 1)}{r - 1}"
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    green = PatternFill("solid", fgColor="C6EFCE")
    ws_g.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["8"], fill=red))
    ws_g.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["4", "8"], fill=yellow))
    ws_g.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["0.01", "4"], fill=green))
    ws_g.column_dimensions["A"].width = 22
    for i in range(2, len(dates_all) + 2):
        ws_g.column_dimensions[get_column_letter(i)].width = 12
    ws_g.freeze_panes = "B4"

    wb.save(path)
    return path, checks


if __name__ == "__main__":
    pass