# -*- coding: utf-8 -*-
"""CI 冒烟测试：公开 APS 标准数据集（data/）排产 + 5-Sheet 公式联动输出。

不依赖任何客户数据；CI 与本地均可运行。
运行: python tests/test_public_data.py
"""
import json
import os
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
for _p in (REPO,):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from aps_engine.api import solve                       # noqa: E402
from aps_engine.audit import audit_result              # noqa: E402
from aps_engine.export_formula import export_formula_xlsx  # noqa: E402
from aps_engine.schema import validate_inputs          # noqa: E402

DATA = os.path.join(REPO, "data")


def main():
    orders = json.load(open(os.path.join(DATA, "orders.json"), encoding="utf-8"))
    lines = json.load(open(os.path.join(DATA, "lines.json"), encoding="utf-8"))
    products = json.load(open(os.path.join(DATA, "products.json"), encoding="utf-8"))

    errs = validate_inputs(orders, lines, products)
    assert not errs, "输入校验失败: " + "\n".join(errs[:10])

    result = solve(orders, lines, products, engine="cp", time_limit=20)
    s = result["summary"]
    assert s["on_time_rate"] == 1.0, f"准时率异常: {s['on_time_rate']}"
    tasks = sum(len(g["tasks"]) for g in result["schedule"])
    assert tasks == len(orders), f"任务数不符: {tasks} != {len(orders)}"

    audited_tasks, issues = audit_result(result)
    assert not issues, "审计发现问题: " + json.dumps(issues, ensure_ascii=False)[:500]

    tmp = tempfile.mkdtemp(prefix="aps_ci_")
    path, checks = export_formula_xlsx(result, products, os.path.join(tmp, "排产结果.xlsx"))
    assert os.path.exists(path) and os.path.getsize(path) > 0

    # 公式联动校验：重新打开，确认 5 个 Sheet 与公式串、逐项 expected 可复算
    from openpyxl import load_workbook
    wb = load_workbook(path)
    need = {"排产明细", "主数据", "汇总", "甘特图", "假设与口径"}
    assert need.issubset(wb.sheetnames), f"缺 Sheet: {need - set(wb.sheetnames)}"
    ws = wb["排产明细"]
    n_formula = 0
    for col in ("J", "K", "L", "M"):
        for rr in range(2, len(checks) // 2 + 2):
            v = ws[f"{col}{rr}"].value
            assert isinstance(v, str) and v.startswith("="), f"{col}{rr} 非公式: {v!r}"
            n_formula += 1
    assert len(checks) >= 2 * len(orders), f"校验项不足: {len(checks)}"
    wb.close()

    print(f"✅ 公开数据集冒烟通过: {tasks} 任务 准时率 {s['on_time_rate']:.0%} "
          f"audit {len(issues)} 问题 公式单元格 {n_formula} 个 校验项 {len(checks)} 条")


if __name__ == "__main__":
    main()
